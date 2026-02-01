#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Course Assistant 任务配置生成器
动态生成不同难度的任务配置，包括：
- Excel 学生名单
- 邮件提交记录
- Evaluation 配置
"""

import json
import random
from pathlib import Path
from typing import List, Dict, Any, Tuple
import argparse


class CourseAssistantConfigGenerator:
    """课程助理任务配置生成器"""
    
    # 英文姓名库 (50 first names × 45 last names = 2250 unique combinations)
    FIRST_NAMES = [
        # Male names
        "James", "John", "Robert", "Michael", "William",
        "David", "Richard", "Joseph", "Thomas", "Christopher",
        "Daniel", "Matthew", "Anthony", "Mark", "Donald",
        "Steven", "Paul", "Andrew", "Joshua", "Kenneth",
        "Kevin", "Brian", "George", "Timothy", "Ronald",
        # Female names
        "Mary", "Patricia", "Jennifer", "Linda", "Elizabeth",
        "Barbara", "Susan", "Jessica", "Sarah", "Karen",
        "Emma", "Olivia", "Ava", "Isabella", "Sophia",
        "Mia", "Charlotte", "Amelia", "Harper", "Evelyn",
        "Abigail", "Emily", "Madison", "Chloe", "Grace"
    ]

    LAST_NAMES = [
        "Smith", "Johnson", "Williams", "Brown", "Jones",
        "Garcia", "Miller", "Davis", "Rodriguez", "Martinez",
        "Hernandez", "Lopez", "Gonzalez", "Wilson", "Anderson",
        "Thomas", "Taylor", "Moore", "Jackson", "Martin",
        "Lee", "Thompson", "White", "Harris", "Sanchez",
        "Clark", "Ramirez", "Lewis", "Robinson", "Walker",
        "Young", "Allen", "King", "Wright", "Scott",
        "Torres", "Nguyen", "Hill", "Flores", "Green",
        "Adams", "Nelson", "Baker", "Hall", "Rivera"
    ]
    
    # NLP 主题内容模板
    NLP_TOPICS = [
        "Recent years have witnessed remarkable progress in Natural Language Processing. Large models like the GPT series have pushed the boundaries of language understanding and generation, paving the way for breakthroughs in multimodal, cross-lingual, and human-machine collaboration domains.",
        "NLP technology is gradually permeating every aspect of life. From intelligent customer service to automatic translation, NLP makes communication between humans and machines more natural. In the future, NLP is expected to achieve deeper semantic understanding.",
        "I believe the future of NLP lies in deep integration with knowledge graphs, reasoning, and other technologies. Only by understanding the knowledge behind language can NLP truly achieve intelligence.",
        "With the development of deep learning, the capabilities of NLP models continue to strengthen. In the future, NLP will focus more on model interpretability and fairness, promoting healthy technological development.",
        "The development of NLP has greatly facilitated information acquisition and knowledge management. In the future, NLP will play a greater role in education, healthcare, and other fields, contributing to social progress.",
        "I am full of expectations for the future of NLP. With the popularization of multilingual models, global information barriers will be further broken down, promoting cultural exchange and understanding.",
        "NLP is not just technology, but a bridge connecting people and the world. In the future, NLP will empower more innovative applications and improve human quality of life.",
        "With the development of pre-trained models and transfer learning, the application threshold of NLP has been greatly lowered. In the future, NLP will become more inclusive, serving a wider range of people.",
        "NLP's progress enables machines to better understand human emotions and intentions. In the future, affective computing and personalized dialogue will become important directions for NLP.",
        "I think the challenge of NLP lies in how to handle complex contexts and implicit semantics. In the future, NLP models will pay more attention to context and reasoning capabilities.",
        "The development of NLP technology has promoted the popularity of applications such as intelligent assistants and automatic summarization. In the future, NLP will show greater potential in cross-domain knowledge integration.",
        "AGI is coming soon. The development of NLP technology has promoted the popularity of applications such as intelligent assistants and automatic summarization. In the future, NLP will show greater potential in cross-domain knowledge integration.",
        "Natural Language Processing represents the intersection of linguistics and artificial intelligence. As transformers revolutionize the field, we're witnessing unprecedented advances in machine understanding of human language.",
        "The evolution of attention mechanisms has fundamentally changed how we approach sequence-to-sequence tasks. Future NLP systems will likely integrate symbolic reasoning with neural approaches.",
        "Transfer learning and few-shot learning are democratizing NLP, allowing smaller organizations to leverage powerful language models. This trend will accelerate innovation across industries."
    ]
    
    def __init__(self, seed: int = 42):
        """初始化生成器"""
        random.seed(seed)
    
    def generate_student_id(self, year_prefix: int = 2000) -> str:
        """生成学号"""
        suffix = random.randint(10000, 99999)
        return f"{year_prefix}{suffix}"
    
    def generate_students(self, num_students: int, dropout_probability: float = 0.1) -> List[Dict[str, Any]]:
        """生成学生列表
        
        Args:
            num_students: 学生总数
            dropout_probability: 退课概率
            
        Returns:
            学生列表，每个学生包含: name, student_id, email, status
        """
        students = []
        used_ids = set()
        used_names = set()
        
        for i in range(num_students):
            # 生成唯一姓名
            while True:
                first_name = random.choice(self.FIRST_NAMES)
                last_name = random.choice(self.LAST_NAMES)
                full_name = f"{first_name} {last_name}"
                if full_name not in used_names:
                    used_names.add(full_name)
                    break
            
            # 生成唯一学号
            while True:
                student_id = self.generate_student_id()
                if student_id not in used_ids:
                    used_ids.add(student_id)
                    break
            
            # 生成邮箱
            email_username = f"{first_name.lower()}{last_name.lower()}{random.randint(1, 99)}"
            email = f"{email_username}@mcp.com"
            
            # 生成密码
            password = ''.join(random.choices('abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789$', k=12))
            
            # 决定是否退课
            status = "dropped" if random.random() < dropout_probability else "enrolled"
            
            students.append({
                "name": full_name,
                "student_id": student_id,
                "email": email,
                "password": password,
                "status": status
            })
        
        return students
    
    def select_submitted_students(self, 
                                   students: List[Dict],
                                   submission_rate: float) -> Tuple[List[Dict], List[Dict]]:
        """选择已提交作业的学生
        
        Args:
            students: 学生列表
            submission_rate: 提交率 (0-1)
            
        Returns:
            (已提交学生列表, 未提交学生列表)
        """
        # 只考虑在册学生
        enrolled_students = [s for s in students if s["status"] == "enrolled"]
        
        # 计算提交人数
        num_submitted = int(len(enrolled_students) * submission_rate)
        
        # 随机选择已提交的学生
        submitted = random.sample(enrolled_students, num_submitted)
        submitted_ids = {s["student_id"] for s in submitted}
        
        # 未提交的学生
        not_submitted = [s for s in enrolled_students if s["student_id"] not in submitted_ids]
        
        return submitted, not_submitted
    
    def generate_email_content(self, student: Dict) -> Dict[str, str]:
        """生成学生的邮件提交内容"""
        content = random.choice(self.NLP_TOPICS)
        
        return {
            "sender_name": student["name"],
            "subject": f"nlp-presentation-{student['student_id']}-{student['name']}",
            "content": f"<html><body><p>{content}</p></body></html>",
            "content_type": "html"
        }
    
    def save_excel_file(self, students: List[Dict], output_path: Path):
        """保存学生信息到 Excel 文件"""
        try:
            from openpyxl import Workbook
        except ImportError:
            print("❌ 错误: openpyxl 未安装，无法创建 Excel 文件")
            print("💡 请运行以下命令安装: pip install openpyxl")
            print("   或者使用 conda: conda install openpyxl")
            raise ImportError("openpyxl is required to create Excel files. Please install it with: pip install openpyxl")
        
        # 创建 Excel 文件
        wb = Workbook()
        ws = wb.active
        ws.title = "NLP Course Students"
        
        # 写入表头
        ws.append(["Name", "Student ID", "Email", "Status"])
        
        # 写入学生数据
        for student in students:
            ws.append([
                student["name"],
                student["student_id"],
                student["email"],
                student["status"]
            ])
        
        # 保存文件
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        print(f"✅ 已保存: {output_path}")
    
    def save_emails_jsonl(self, submitted_students: List[Dict], output_path: Path):
        """保存邮件提交记录到 JSONL 文件"""
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            for student in submitted_students:
                email_data = self.generate_email_content(student)
                f.write(json.dumps(email_data, ensure_ascii=False) + '\n')
        
        print(f"✅ 已保存: {output_path} ({len(submitted_students)} 封邮件)")
    
    def save_students_info(self, students: List[Dict], output_path: Path):
        """保存完整的学生信息（包括密码）到 JSON 文件"""
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(students, f, ensure_ascii=False, indent=2)
        
        print(f"✅ 已保存学生信息: {output_path} ({len(students)} 个学生)")
    
    def save_evaluation_config(self, 
                               not_submitted_students: List[Dict],
                               num_check_students: int,
                               output_dir: Path):
        """保存 evaluation 配置
        
        新版本的评估脚本会自动从 Excel 和 emails.jsonl 读取所有学生数据，
        因此不需要预先生成学生列表。
        
        重要说明：
        - 新版本会检查所有未提交作业的在册学生
        - 已退课学生和已提交作业的学生会自动排除
        """
        print(f"\n   未提交学生总数: {len(not_submitted_students)} 人")
        print(f"   ✅ 新版本评估脚本会检查所有未提交学生")
        
        # 更新 evaluation check_local.py 文件
        eval_file = output_dir / "evaluation" / "check_local.py"
        
        # 读取模板
        template_path = Path(__file__).parent / "evaluation_template.py"
        if template_path.exists():
            with open(template_path, 'r', encoding='utf-8') as f:
                template = f.read()
        else:
            # 如果没有模板，使用默认模板
            template = self.get_evaluation_template()
        
        eval_file.parent.mkdir(parents=True, exist_ok=True)
        with open(eval_file, 'w', encoding='utf-8') as f:
            f.write(template)
        
        print(f"✅ 已更新: {eval_file}")
        print(f"   将检查所有 {len(not_submitted_students)} 个未提交学生")
        
        return not_submitted_students
    
    def get_evaluation_template(self) -> str:
        """返回 evaluation 脚本模板"""
        return '''#!/usr/bin/env python3
"""
邮件内容检查脚本 - Course Assistant 任务评估

检查目标：
- 检查未提交作业的在册学生是否收到催促邮件
- 邮件主题必须为 "nlp-course-emergency"
- 邮件内容必须包含学生的姓名和学号

评估标准：
1. ✅ 每个学生收到恰好 1 封符合要求的邮件
2. ✅ 邮件主题正确
3. ✅ 邮件内容包含学生姓名和学号
4. ❌ 不应有主题相同但内容不符的多余邮件
"""

import os
import sys
import re
from pathlib import Path
from typing import List, Tuple, Dict



from mcp_convert.mcps.email.database_utils import EmailDatabase


def extract_email_body(email_dict: Dict) -> str:
    """从邮件字典中提取正文（优先 body，fallback 到 html_body 并去除标签）"""
    # 优先使用纯文本 body
    body = email_dict.get('body', '')
    if body:
        return body
    
    # Fallback 到 html_body 并去除 HTML 标签
    html_body = email_dict.get('html_body', '')
    if html_body:
        # 简单去除 HTML 标签
        clean_body = re.sub('<[^<]+?>', '', html_body)
        return clean_body
    
    return ''


def check_account_emails_db(db: EmailDatabase,
                            email_address: str,
                            password: str,
                            required_keywords: List[str],
                            account_label: str) -> Tuple[bool, Dict]:
    """检查指定账户的 nlp-course-emergency 邮件（使用数据库）"""
    passed = True
    valid_mail_info = None
    
    try:
        # 登录用户
        try:
            db.login(email_address, password)
        except ValueError as e:
            print(f"❌ [{account_label}] 登录失败: {e}")
            return False, None
        
        # 搜索主题为 nlp-course-emergency 的邮件
        search_result = db.search_emails(query="nlp-course-emergency", folder="INBOX", page=1, page_size=100)
        emails = search_result.get('emails', [])
        
        if not emails:
            print(f"❌ [{account_label}] 没有找到主题为 nlp-course-emergency 的邮件")
            db.logout()
            return False, None
        
        valid_count = 0
        extra_msgs = []
        
        for email_data in emails:
            subject = email_data.get('subject', 'Unknown Subject')
            sender = email_data.get('from', 'Unknown Sender')
            body = extract_email_body(email_data)
            
            # 检查所有关键词
            if all(kw in body for kw in required_keywords):
                valid_count += 1
                valid_mail_info = {
                    'account': account_label,
                    'subject': subject,
                    'sender': sender,
                    'body': body
                }
            else:
                snippet = body[:60].replace('\\n', ' ').replace('\\r', ' ')
                extra_msgs.append(f"主题: {subject} | 发件人: {sender} | 正文片段: {snippet}")
        
        # 验证结果
        if valid_count == 0:
            print(f"❌ [{account_label}] 没有找到正文包含所有关键词({required_keywords})的邮件")
            passed = False
        elif valid_count > 1:
            print(f"❌ [{account_label}] 找到{valid_count}封正文包含所有关键词({required_keywords})的邮件，应只有1封")
            passed = False
        
        if extra_msgs:
            print(f"❌ [{account_label}] 存在{len(extra_msgs)}封主题为 nlp-course-emergency 但正文不符的多余邮件:")
            for msg in extra_msgs:
                print(f"   • {msg}")
            passed = False
        
        if passed:
            print(f"✅ [{account_label}] 邮件检查通过")
        
        db.logout()
        
    except Exception as e:
        print(f"❌ [{account_label}] 检查过程中发生异常: {e}")
        import traceback
        traceback.print_exc()
        passed = False
    
    return passed, valid_mail_info


def load_students_from_config(config_dir: Path) -> Tuple[List[Dict], List[Dict], List[Dict]]:
    """从配置文件加载学生信息
    
    Returns:
        (未提交的在册学生, 已提交的学生, 已退课的学生)
    """
    import json
    
    # 读取 initial_workspace 中的 Excel 文件
    excel_path = config_dir / "initial_workspace" / "nlp_statistics.xlsx"
    if not excel_path.exists():
        print(f"❌ Excel 文件不存在: {excel_path}")
        return [], [], []
    
    # 读取 files 中的 emails.jsonl 获取已提交学生
    emails_jsonl = config_dir / "files" / "emails.jsonl"
    submitted_student_ids = set()
    if emails_jsonl.exists():
        with open(emails_jsonl, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    email_data = json.loads(line)
                    # 从主题中提取学号: nlp-presentation-{student_id}-{name}
                    subject = email_data.get('subject', '')
                    import re
                    match = re.search(r'nlp-presentation-(\\d+)-', subject)
                    if match:
                        submitted_student_ids.add(match.group(1))
                except:
                    continue
    
    # 读取 Excel 文件
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("❌ 错误: openpyxl 未安装")
        return [], [], []
    
    wb = load_workbook(excel_path)
    ws = wb.active
    
    all_students = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过表头
        if not row[0]:  # 如果姓名为空，跳过
            continue
        student = {
            'name': row[0],
            'student_id': str(row[1]),
            'email': row[2],
            'status': row[3]
        }
        all_students.append(student)
    
    # 分类学生
    not_submitted_enrolled = []  # 未提交的在册学生
    submitted_students = []       # 已提交的学生
    dropped_students = []         # 已退课的学生
    
    for student in all_students:
        if student['status'] == 'dropped':
            dropped_students.append(student)
        elif student['student_id'] in submitted_student_ids:
            submitted_students.append(student)
        else:
            not_submitted_enrolled.append(student)
    
    return not_submitted_enrolled, submitted_students, dropped_students


def main():
    """
    评估函数 - 检查未提交作业的在册学生是否收到催促邮件
    
    检查逻辑：
    1. 检查所有未提交作业的在册学生都收到催促邮件
    2. 邮件主题必须为 "nlp-course-emergency"
    3. 邮件内容必须包含学生的姓名和学号
    4. 确保已提交作业的学生和已退课的学生没有收到邮件
    
    注意：
    - 已退课学生（status="dropped"）不应收到邮件
    - 已提交作业的学生不应收到邮件
    """
    
    # 初始化 EmailDatabase
    # 尝试从环境变量获取数据库目录
    email_db_dir = os.environ.get('EMAIL_DATA_DIR')
    if not email_db_dir:
        # Fallback 到默认位置
        email_db_dir = str(MCP_CONVERT_PATH / "mcps" / "email" / "data")
    
    # 加载学生配置
    not_submitted, submitted, dropped = load_students_from_config(task_dir)
    
    print(f"\\n📊 学生统计:")
    print(f"   ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print(f"   未提交作业的在册学生: {len(not_submitted)} 人 (应收到邮件)")
    print(f"   已提交作业的学生: {len(submitted)} 人 (不应收到邮件)")
    print(f"   已退课的学生: {len(dropped)} 人 (不应收到邮件)")
    print(f"   ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\\n")
    
    if not not_submitted:
        print("❌ 错误: 没有找到未提交作业的在册学生")
        return 0
    
    # 检查所有未提交的在册学生
    all_passed = True
    valid_mails = []
    
    print("=" * 60)
    print("🔍 检查未提交作业的在册学生是否收到催促邮件...")
    print("=" * 60)
    
    for student in not_submitted:
        student_name = student['name']
        student_email = student['email']
        student_id = student['student_id']
        
        # 由于我们没有存储所有学生的密码，这里需要从数据库的 users 中获取
        user_info = db.users.get(student_email)
        if not user_info:
            print(f"\\n❌ 学生 {student_name} ({student_email}) 在数据库中不存在")
            all_passed = False
            continue
        
        password = user_info.get('password', '')
        
        print(f"\\n📧 检查学生 {student_name} 的收件箱: {student_email}")
        print(f"🔍 检查学生 {student_name} 是否收到催促邮件...")
        
        passed, valid_mail_info = check_account_emails_db(
            db,
            student_email,
            password,
            [student_name, student_id],
            student_name
        )
        
        if valid_mail_info:
            valid_mails.append(valid_mail_info)
        if not passed:
            all_passed = False
    
    print("\\n" + "=" * 60)
    print(f"📊 检查结果汇总")
    print("=" * 60)
    print(f"   应收到邮件的学生数: {len(not_submitted)}")
    print(f"   实际收到邮件的学生数: {len(valid_mails)}")
    print("=" * 60 + "\\n")
    
    if all_passed:
        print("\\n🎉 所有账户邮件检查通过！\\n")
        print("====== 合格邮件内容 ======")
        for mail in valid_mails:
            print(f"账户: {mail['account']}")
            print(f"发件人: {mail['sender']}")
            print(f"主题: {mail['subject']}")
            print(f"正文:\\n{mail['body']}\\n")
            print("------------------------")
        print("========================\\n")
    else:
        print("\\n💥 邮件检查未通过！")
        print("⚠️  以下学生应该收到邮件但未通过检查:")
        for student in not_submitted:
            found = any(mail['account'] == student['name'] for mail in valid_mails)
            if not found:
                print(f"   • {student['name']} ({student['email']})")
    
    return 1 if all_passed else 0

if __name__ == '__main__':
    exit(main())
'''
    
    def generate_config(self,
                       output_dir: Path,
                       num_students: int = 15,
                       dropout_rate: float = 0.1,
                       submission_rate: float = 0.7,
                       num_check_students: int = 2,
                       seed: int = None):
        """生成完整的任务配置
        
        Args:
            output_dir: 输出目录
            num_students: 学生总数
            dropout_rate: 退课率 (0-1)
            submission_rate: 提交率 (0-1)
            num_check_students: 需要检查的学生数量
            seed: 随机种子
        """
        if seed is not None:
            random.seed(seed)
        
        print(f"🎲 生成课程助理任务配置...")
        print(f"   学生总数: {num_students}")
        print(f"   退课率: {dropout_rate:.0%}")
        print(f"   提交率: {submission_rate:.0%}")
        print(f"   检查学生数: {num_check_students}")
        
        # 1. 生成学生列表
        print(f"\n📝 生成学生名单...")
        students = self.generate_students(num_students, dropout_rate)
        
        enrolled_students = [s for s in students if s["status"] == "enrolled"]
        dropped_students = [s for s in students if s["status"] == "dropped"]
        
        print(f"   总学生数: {num_students}")
        print(f"   在册学生: {len(enrolled_students)}")
        print(f"   退课学生: {len(dropped_students)}")
        
        # 2. 选择已提交的学生
        print(f"\n📧 生成邮件提交记录...")
        submitted, not_submitted = self.select_submitted_students(students, submission_rate)
        
        print(f"   已提交: {len(submitted)}")
        print(f"   未提交: {len(not_submitted)}")
        
        # 3. 保存 Excel 文件
        excel_path = output_dir / "initial_workspace" / "nlp_statistics.xlsx"
        self.save_excel_file(students, excel_path)
        
        # 4. 保存学生信息（包括密码）到 JSON
        students_info_path = output_dir / "files" / "students_info.json"
        self.save_students_info(students, students_info_path)
        
        # 5. 保存邮件 JSONL 文件
        emails_path = output_dir / "files" / "emails.jsonl"
        self.save_emails_jsonl(submitted, emails_path)
        
        # 6. 保存 evaluation 配置
        print(f"\n🔍 生成 evaluation 配置...")
        check_students = self.save_evaluation_config(not_submitted, num_check_students, output_dir)
        
        # 7. 统计信息
        print(f"\n📊 任务统计:")
        print(f"   ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        print(f"   总学生数: {num_students}")
        print(f"   ├─ 在册学生: {len(enrolled_students)}")
        print(f"   │  ├─ 已提交作业: {len(submitted)} (不需要催促)")
        print(f"   │  └─ 未提交作业: {len(not_submitted)} (需要催促)")
        print(f"   └─ 退课学生: {len(dropped_students)} (不需要催促)")
        print(f"   ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        print(f"   \n   🎯 Evaluation 将检查所有 {len(check_students)} 个未提交学生")
        
        print(f"\n✅ 配置生成完成！")
        
        return {
            "total_students": num_students,
            "enrolled": len(enrolled_students),
            "dropped": len(dropped_students),
            "submitted": len(submitted),
            "not_submitted": len(not_submitted),
            "to_remind": len(check_students)
        }


def main():
    parser = argparse.ArgumentParser(description="Course Assistant 任务配置生成器")
    
    # 基本参数
    parser.add_argument("--num-students", type=int, default=15,
                       help="学生总数 (默认: 15)")
    parser.add_argument("--dropout-rate", type=float, default=0.1,
                       help="退课率 (0-1, 默认: 0.1)")
    parser.add_argument("--submission-rate", type=float, default=0.7,
                       help="作业提交率 (0-1, 默认: 0.7)")
    parser.add_argument("--num-check", type=int, default=2,
                       help="需要检查的学生数量 (已废弃，现在会检查所有未提交的学生，保留此参数仅为向后兼容)")
    
    # 其他参数
    parser.add_argument("--seed", type=int, default=42,
                       help="随机种子 (默认: 42)")
    parser.add_argument("--output-dir", type=str, default=".",
                       help="输出目录 (默认: 当前目录)")
    
    args = parser.parse_args()
    
    # 生成配置
    generator = CourseAssistantConfigGenerator(seed=args.seed)
    output_dir = Path(args.output_dir)
    
    stats = generator.generate_config(
        output_dir=output_dir,
        num_students=args.num_students,
        dropout_rate=args.dropout_rate,
        submission_rate=args.submission_rate,
        num_check_students=args.num_check,
        seed=args.seed
    )
    
    print(f"\n💡 使用示例:")
    print(f"   python preprocess/main.py --agent_workspace /path/to/workspace")
    print(f"\n🎉 任务配置已生成！")


if __name__ == "__main__":
    main()

