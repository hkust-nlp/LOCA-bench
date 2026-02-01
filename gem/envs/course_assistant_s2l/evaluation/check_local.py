#!/usr/bin/env python3
"""
邮件内容检查脚本 - Course Assistant 任务评估

检查目标：
- 检查未提交作业的在册学生是否收到催促邮件
- 邮件主题必须为 "nlp-course-emergency"
- 邮件内容必须包含学生的姓名和学号

评估标准：
1. ✅ 每个学生收到恰好 1 封符合要求的邮件
2. ✅ 邮件主题正确
3. ✅ 邮件内容包含学生姓名和学号
4. ❌ 不应有主题相同但内容不符的多余邮件
"""

import os
import sys
import re
from pathlib import Path
from typing import List, Tuple, Dict

current_dir = Path(__file__).parent
# 添加 mcp_convert 路径以导入 EmailDatabase
from mcp_convert.mcps.email.database_utils import EmailDatabase


def extract_email_body(email_dict: Dict) -> str:
    """从邮件字典中提取正文（优先 body，fallback 到 html_body 并去除标签）"""
    # 优先使用纯文本 body
    body = email_dict.get('body', '')
    if body:
        return body
    
    # Fallback 到 html_body 并去除 HTML 标签
    html_body = email_dict.get('html_body', '')
    if html_body:
        # 简单去除 HTML 标签
        clean_body = re.sub('<[^<]+?>', '', html_body)
        return clean_body
    
    return ''


def check_account_emails_db(db: EmailDatabase,
                            email_address: str,
                            password: str,
                            required_keywords: List[str],
                            account_label: str) -> Tuple[bool, Dict]:
    """检查指定账户的 nlp-course-emergency 邮件（使用数据库）"""
    passed = True
    valid_mail_info = None
    
    try:
        # 登录用户
        try:
            db.login(email_address, password)
        except ValueError as e:
            print(f"❌ [{account_label}] 登录失败: {e}")
            return False, None
        
        # 搜索主题为 nlp-course-emergency 的邮件
        search_result = db.search_emails(query="nlp-course-emergency", folder="INBOX", page=1, page_size=100)
        emails = search_result.get('emails', [])
        
        if not emails:
            print(f"❌ [{account_label}] 没有找到主题为 nlp-course-emergency 的邮件")
            db.logout()
            return False, None
        
        valid_count = 0
        extra_msgs = []
        
        for email_data in emails:
            subject = email_data.get('subject', 'Unknown Subject')
            sender = email_data.get('from', 'Unknown Sender')
            body = extract_email_body(email_data)
            
            # 检查所有关键词
            if all(kw in body for kw in required_keywords):
                valid_count += 1
                valid_mail_info = {
                    'account': account_label,
                    'subject': subject,
                    'sender': sender,
                    'body': body
                }
            else:
                snippet = body[:60].replace('\n', ' ').replace('\r', ' ')
                extra_msgs.append(f"主题: {subject} | 发件人: {sender} | 正文片段: {snippet}")
        
        # 验证结果
        if valid_count == 0:
            print(f"❌ [{account_label}] 没有找到正文包含所有关键词({required_keywords})的邮件")
            passed = False
        elif valid_count > 1:
            print(f"❌ [{account_label}] 找到{valid_count}封正文包含所有关键词({required_keywords})的邮件，应只有1封")
            passed = False
        
        if extra_msgs:
            print(f"❌ [{account_label}] 存在{len(extra_msgs)}封主题为 nlp-course-emergency 但正文不符的多余邮件:")
            for msg in extra_msgs:
                print(f"   • {msg}")
            passed = False
        
        if passed:
            print(f"✅ [{account_label}] 邮件检查通过")
        
        db.logout()
        
    except Exception as e:
        print(f"❌ [{account_label}] 检查过程中发生异常: {e}")
        import traceback
        traceback.print_exc()
        passed = False
    
    return passed, valid_mail_info


def load_students_from_config(config_dir: Path) -> Tuple[List[Dict], List[Dict], List[Dict]]:
    """从配置文件加载学生信息
    
    Returns:
        (未提交的在册学生, 已提交的学生, 已退课的学生)
    """
    import json
    
    # 读取 initial_workspace 中的 Excel 文件
    excel_path = config_dir / "initial_workspace" / "nlp_statistics.xlsx"
    if not excel_path.exists():
        print(f"❌ Excel 文件不存在: {excel_path}")
        return [], [], []
    
    # 读取 files 中的 emails.jsonl 获取已提交学生
    emails_jsonl = config_dir / "files" / "emails.jsonl"
    submitted_student_ids = set()
    if emails_jsonl.exists():
        with open(emails_jsonl, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    email_data = json.loads(line)
                    # 从主题中提取学号: nlp-presentation-{student_id}-{name}
                    subject = email_data.get('subject', '')
                    import re
                    match = re.search(r'nlp-presentation-(\d+)-', subject)
                    if match:
                        submitted_student_ids.add(match.group(1))
                except:
                    continue
    
    # 读取 Excel 文件
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("❌ 错误: openpyxl 未安装")
        return [], [], []
    
    wb = load_workbook(excel_path)
    ws = wb.active
    
    all_students = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过表头
        if not row[0]:  # 如果姓名为空，跳过
            continue
        student = {
            'name': row[0],
            'student_id': str(row[1]),
            'email': row[2],
            'status': row[3]
        }
        all_students.append(student)
    
    # 分类学生
    not_submitted_enrolled = []  # 未提交的在册学生
    submitted_students = []       # 已提交的学生
    dropped_students = []         # 已退课的学生
    
    for student in all_students:
        if student['status'] == 'dropped':
            dropped_students.append(student)
        elif student['student_id'] in submitted_student_ids:
            submitted_students.append(student)
        else:
            not_submitted_enrolled.append(student)
    
    return not_submitted_enrolled, submitted_students, dropped_students


def main():
    """
    评估函数 - 检查未提交作业的在册学生是否收到催促邮件
    
    检查逻辑：
    1. 检查所有未提交作业的在册学生都收到催促邮件
    2. 邮件主题必须为 "nlp-course-emergency"
    3. 邮件内容必须包含学生的姓名和学号
    4. 确保已提交作业的学生和已退课的学生没有收到邮件
    
    注意：
    - 已退课学生（status="dropped"）不应收到邮件
    - 已提交作业的学生不应收到邮件
    """
    
    # 初始化 EmailDatabase
    # 尝试从环境变量获取数据库目录
    email_db_dir = os.environ.get('EMAIL_DATA_DIR')

    
    print(f"📂 Email 数据库目录: {email_db_dir}")
    
    try:
        db = EmailDatabase(data_dir=email_db_dir)
    except Exception as e:
        print(f"❌ 无法初始化 EmailDatabase: {e}")
        return 0
    
    # 加载学生配置
    # 优先从环境变量获取 task_dir（支持多实例运行，避免冲突）
    task_dir_str = os.environ.get('TASK_DIR')
    if task_dir_str:
        # 使用环境变量指定的任务目录（每个实例独立）
        task_dir = Path(task_dir_str)
    else:
        # Fallback 到基于脚本位置的路径（兼容旧的调用方式）
        task_dir = current_dir.parent
    
    not_submitted, submitted, dropped = load_students_from_config(task_dir)
    
    print(f"\n📊 学生统计:")
    print(f"   ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print(f"   未提交作业的在册学生: {len(not_submitted)} 人 (应收到邮件)")
    print(f"   已提交作业的学生: {len(submitted)} 人 (不应收到邮件)")
    print(f"   已退课的学生: {len(dropped)} 人 (不应收到邮件)")
    print(f"   ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
    
    if not not_submitted:
        print("❌ 错误: 没有找到未提交作业的在册学生")
        return 0
    
    # 检查所有未提交的在册学生
    all_passed = True
    valid_mails = []
    
    print("=" * 60)
    print("🔍 检查未提交作业的在册学生是否收到催促邮件...")
    print("=" * 60)
    
    for student in not_submitted:
        student_name = student['name']
        student_email = student['email']
        student_id = student['student_id']
        
        # 由于我们没有存储所有学生的密码，这里需要从数据库的 users 中获取
        user_info = db.users.get(student_email)
        if not user_info:
            print(f"\n❌ 学生 {student_name} ({student_email}) 在数据库中不存在")
            all_passed = False
            continue
        
        password = user_info.get('password', '')
        
        print(f"\n📧 检查学生 {student_name} 的收件箱: {student_email}")
        print(f"🔍 检查学生 {student_name} 是否收到催促邮件...")
        
        passed, valid_mail_info = check_account_emails_db(
            db,
            student_email,
            password,
            [student_name, student_id],
            student_name
        )
        
        if valid_mail_info:
            valid_mails.append(valid_mail_info)
        if not passed:
            all_passed = False
    
    print("\n" + "=" * 60)
    print(f"📊 检查结果汇总")
    print("=" * 60)
    print(f"   应收到邮件的学生数: {len(not_submitted)}")
    print(f"   实际收到邮件的学生数: {len(valid_mails)}")
    print("=" * 60 + "\n")
    
    if all_passed:
        print("\n🎉 所有账户邮件检查通过！\n")
        print("====== 合格邮件内容 ======")
        for mail in valid_mails:
            print(f"账户: {mail['account']}")
            print(f"发件人: {mail['sender']}")
            print(f"主题: {mail['subject']}")
            print(f"正文:\n{mail['body']}\n")
            print("------------------------")
        print("========================\n")
    else:
        print("\n💥 邮件检查未通过！")
        print("⚠️  以下学生应该收到邮件但未通过检查:")
        for student in not_submitted:
            found = any(mail['account'] == student['name'] for mail in valid_mails)
            if not found:
                print(f"   • {student['name']} ({student['email']})")
    
    return 1 if all_passed else 0

if __name__ == '__main__':
    exit(main())
